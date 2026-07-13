from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import re
from typing import Iterable

from openpyxl import load_workbook
import pandas as pd
from openpyxl.utils.exceptions import InvalidFileException

from app.config import REQUIRED_COLUMNS


MONTH_NAME_MAP = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

MONTH_LABELS = {month: f"{month}月" for month in range(1, 13)}

TARGET_YEAR_CANDIDATES = ["Year", "年份", "年度", "目标年度", "Calendar Year"]
TARGET_MONTH_CANDIDATES = ["Month", "月份", "月", "月份序号"]
TARGET_ORIGINAL_CANDIDATES = [
    "Target",
    "Sales Target",
    "Original Target",
    "Original Sales Target",
    "目标",
    "销售目标",
    "原始目标",
    "原始销售目标",
    "月度目标",
]
TARGET_REVISED_CANDIDATES = [
    "Revised Target",
    "Revised Sales Target",
    "Adjusted Target",
    "调整后目标",
    "调整销售目标",
    "修订目标",
]
TARGET_ANNUAL_CANDIDATES = [
    "Annual Target",
    "Year Target",
    "Annual Sales Target",
    "年度目标",
    "年度合计",
    "全年目标",
    "全年合计",
]
TARGET_NOTES_CANDIDATES = ["Notes", "Note", "备注", "说明"]


@dataclass(frozen=True)
class TargetSheetCandidate:
    sheet_name: str
    header_row: int
    layout: str
    score: int
    year_column: str | None
    month_column: str | None
    original_target_column: str | None
    revised_target_column: str | None
    annual_target_column: str | None
    notes_column: str | None
    month_columns: dict[int, str]

    @property
    def label(self) -> str:
        layout_label = "长表" if self.layout == "long" else "横向月份"
        return f"{self.sheet_name}（{layout_label}，第 {self.header_row + 1} 行为表头）"


@dataclass(frozen=True)
class TargetWorkbookAnalysis:
    candidates: list[TargetSheetCandidate]
    sheet_names: list[str]


@dataclass(frozen=True)
class XFTargetWorkbook:
    amount_data: pd.DataFrame
    case_data: pd.DataFrame
    company_targets: pd.DataFrame
    annual_targets: dict[int, float]
    amount_sheet: str | None
    case_sheet: str | None
    target_year: int | None
    product_group_count: int
    company_annual_amount_target: float | None
    company_annual_case_target: float | None
    structure_label: str


def _normalize_name(value: object) -> str:
    text = str(value).strip().lower()
    for token in [" ", "_", "-", ".", "/", "\\", "（", "）", "(", ")"]:
        text = text.replace(token, "")
    return text


def _first_existing(columns: Iterable[object], candidates: list[str]) -> str | None:
    normalized = {_normalize_name(col): str(col).strip() for col in columns}
    for candidate in candidates:
        match = normalized.get(_normalize_name(candidate))
        if match:
            return match
    return None


def parse_month(value: object) -> int | None:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return int(value.month)
    text = str(value).strip()
    if not text:
        return None
    lowered = text.lower().replace("月", "").replace(".", "").strip()
    if lowered in MONTH_NAME_MAP:
        return MONTH_NAME_MAP[lowered]
    try:
        month = int(float(lowered))
    except ValueError:
        return None
    return month if 1 <= month <= 12 else None


def _to_number(value: object) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip()
    if not text or text in {"——", "-", "#DIV/0!", "#VALUE!", "#REF!", "#N/A"}:
        return None
    text = text.replace(",", "").replace("£", "").replace("%", "")
    try:
        return float(text)
    except ValueError:
        return None


def _year_from_text(*values: object) -> int | None:
    for value in values:
        if value is None:
            continue
        match = re.search(r"(20\d{2})", str(value))
        if match:
            return int(match.group(1))
    return None


def _open_workbook(excel_file, data_only: bool = True):
    try:
        if isinstance(excel_file, (bytes, bytearray)):
            return load_workbook(BytesIO(excel_file), data_only=data_only)
        if hasattr(excel_file, "seek"):
            excel_file.seek(0)
        return load_workbook(excel_file, data_only=data_only)
    except (InvalidFileException, OSError, ValueError) as exc:
        raise ValueError("目标 Excel 无法读取，请确认文件是有效的 .xlsx 工作簿。") from exc


def _filled_header_values(ws, row: int) -> dict[int, object]:
    values: dict[int, object] = {}
    current = None
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if value is not None:
            current = value
        values[col] = current
    return values


def parse_xf_target_amount_sheet(excel_file, sheet_name: str = "2026销售目标金额") -> pd.DataFrame:
    """Parse XF's fixed multi-header amount target template into a long table."""
    wb = _open_workbook(excel_file, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"未找到金额目标工作表：{sheet_name}")
    ws = wb[sheet_name]
    month_headers = _filled_header_values(ws, 1)
    target_year = _year_from_text(sheet_name, ws.cell(1, 2).value)

    month_columns: dict[int, dict[str, int]] = {}
    annual_plan_col = None
    for col in range(1, ws.max_column + 1):
        group_label = str(month_headers.get(col) or "").strip()
        sub_label = ws.cell(3, col).value
        month = parse_month(group_label)
        if month:
            if sub_label == 2025:
                month_columns.setdefault(month, {})["previous"] = col
            elif sub_label == 2026:
                month_columns.setdefault(month, {})["target"] = col
        if _normalize_name(group_label) == _normalize_name("2026计划"):
            annual_plan_col = col

    if not month_columns:
        raise ValueError("金额目标工作表未识别到 1月-12月的 2025/2026 月度列。")

    rows: list[dict[str, object]] = []
    for row_idx in range(4, ws.max_row + 1):
        first_col = ws.cell(row_idx, 1).value
        product_group = ws.cell(row_idx, 2).value
        is_total = str(first_col).strip().upper() in {"总计", "TOTAL"}
        if is_total:
            product_group_text = "公司整体"
        elif product_group is None or str(product_group).strip() == "":
            continue
        else:
            product_group_text = str(product_group).strip()

        annual_original = _to_number(ws.cell(row_idx, annual_plan_col).value) if annual_plan_col else None
        row_has_value = annual_original is not None
        for month, cols in sorted(month_columns.items()):
            previous = _to_number(ws.cell(row_idx, cols.get("previous", 0)).value) if cols.get("previous") else None
            original = _to_number(ws.cell(row_idx, cols.get("target", 0)).value) if cols.get("target") else None
            if previous is not None or original is not None:
                row_has_value = True
            rows.append(
                {
                    "Year": target_year,
                    "Month": month,
                    "Month Label": MONTH_LABELS[month],
                    "Product Group": product_group_text,
                    "Previous Year Actual": previous,
                    "Original Target": original,
                    "Revised Target": original,
                    "Annual Original Target": annual_original,
                    "Notes": "",
                }
            )
        if not row_has_value and not is_total:
            rows = rows[: -len(month_columns)]

    df = pd.DataFrame.from_records(rows)
    if df.empty:
        raise ValueError("金额目标工作表未提取到有效目标数据。")
    df["Year"] = df["Year"].fillna(target_year).astype("Int64")
    return df


def parse_xf_monthly_case_targets(excel_file, sheet_name: str = "月度箱数需求") -> pd.DataFrame:
    """Parse XF's monthly case target sheet into a long table."""
    wb = _open_workbook(excel_file, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"未找到数量目标工作表：{sheet_name}")
    ws = wb[sheet_name]
    target_year = _year_from_text(sheet_name, ws.cell(1, 1).value)

    header_row = None
    for row in range(1, min(ws.max_row, 12) + 1):
        first_value = _normalize_name(ws.cell(row, 1).value or "")
        if first_value in {_normalize_name("品类"), _normalize_name("Product Group")}:
            header_row = row
            break
    if header_row is None:
        raise ValueError("数量目标工作表未找到“品类 / Product Group”表头。")

    month_cols: dict[int, int] = {}
    annual_col = None
    price_col = None
    for col in range(1, ws.max_column + 1):
        header = ws.cell(header_row, col).value
        month = parse_month(header)
        if month:
            month_cols[month] = col
        normalized = _normalize_name(header or "")
        if "箱价" in str(header or "") or normalized in {"caseprice", "price"}:
            price_col = col
        if normalized in {_normalize_name("全年箱数"), _normalize_name("Annual Case Target")}:
            annual_col = col

    if len(month_cols) < 12:
        raise ValueError("数量目标工作表未识别到完整的 1月-12月箱数列。")

    rows: list[dict[str, object]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        product_group = ws.cell(row_idx, 1).value
        if product_group is None or str(product_group).strip() == "":
            continue
        product_text = str(product_group).strip()
        if product_text == "关键结果" or "全年约需要销售" in product_text:
            break
        if product_text.upper() == "TOTAL":
            product_text = "公司整体"
        annual_case = _to_number(ws.cell(row_idx, annual_col).value) if annual_col else None
        case_price = _to_number(ws.cell(row_idx, price_col).value) if price_col else None
        has_value = annual_case is not None
        pending_rows = []
        for month, col in sorted(month_cols.items()):
            case_target = _to_number(ws.cell(row_idx, col).value)
            if case_target is not None:
                has_value = True
            pending_rows.append(
                {
                    "Year": target_year,
                    "Month": month,
                    "Month Label": MONTH_LABELS[month],
                    "Product Group": product_text,
                    "Case Target": case_target,
                    "Annual Case Target": annual_case,
                    "Case Price": case_price,
                }
            )
        if has_value:
            rows.extend(pending_rows)

    df = pd.DataFrame.from_records(rows)
    if df.empty:
        raise ValueError("数量目标工作表未提取到有效目标数据。")
    df["Year"] = df["Year"].fillna(target_year).astype("Int64")
    return df


def parse_xf_target_workbook(excel_file) -> XFTargetWorkbook:
    amount_sheet = "2026销售目标金额"
    case_sheet = "月度箱数需求"
    amount_df = parse_xf_target_amount_sheet(excel_file, amount_sheet)
    case_df = parse_xf_monthly_case_targets(excel_file, case_sheet)

    year_values = amount_df["Year"].dropna().astype(int).unique().tolist()
    target_year = int(year_values[0]) if year_values else None
    company_targets = (
        amount_df[amount_df["Product Group"].eq("公司整体")]
        .copy()
        .rename(columns={"Previous Year Actual": "Previous Year Target Template Actual"})
    )
    if company_targets.empty:
        raise ValueError("金额目标工作表未识别到“总计”公司整体目标行。")
    company_targets = company_targets[
        ["Year", "Month", "Month Label", "Original Target", "Revised Target", "Notes"]
    ].copy()
    annual_targets: dict[int, float] = {}
    for year, group in company_targets.groupby("Year"):
        annual_targets[int(year)] = float(pd.to_numeric(group["Original Target"], errors="coerce").fillna(0).sum())

    product_groups = amount_df.loc[~amount_df["Product Group"].eq("公司整体"), "Product Group"].dropna().unique().tolist()
    company_annual_amount = annual_targets.get(target_year) if target_year is not None else None
    company_case_rows = case_df[case_df["Product Group"].eq("公司整体")]
    company_annual_case = None
    if not company_case_rows.empty:
        company_annual_case = _to_number(company_case_rows["Annual Case Target"].dropna().iloc[0]) if not company_case_rows["Annual Case Target"].dropna().empty else None

    return XFTargetWorkbook(
        amount_data=amount_df,
        case_data=case_df,
        company_targets=company_targets,
        annual_targets=annual_targets,
        amount_sheet=amount_sheet,
        case_sheet=case_sheet,
        target_year=target_year,
        product_group_count=len(product_groups),
        company_annual_amount_target=company_annual_amount,
        company_annual_case_target=company_annual_case,
        structure_label=f"{amount_sheet} + {case_sheet}",
    )



def _month_column_number(column: object) -> int | None:
    text = str(column).strip()
    normalized = _normalize_name(text).replace("target", "").replace("目标", "")
    if normalized in MONTH_NAME_MAP:
        return MONTH_NAME_MAP[normalized]
    if text.endswith("月"):
        return parse_month(text)
    if normalized.startswith("m") and normalized[1:].isdigit():
        return parse_month(normalized[1:])
    return parse_month(normalized)


def _read_sheet(excel_file, sheet_name: str, header_row: int) -> pd.DataFrame | None:
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, header=header_row)
    except Exception:
        return None
    df = df.dropna(how="all").copy()
    if df.empty:
        return None
    df.columns = [str(col).strip() for col in df.columns]
    df = df.loc[:, ~pd.Index(df.columns).astype(str).str.startswith("Unnamed")]
    return df if not df.empty else None


def _score_candidate(df: pd.DataFrame, sheet_name: str, header_row: int) -> TargetSheetCandidate | None:
    year_col = _first_existing(df.columns, TARGET_YEAR_CANDIDATES)
    month_col = _first_existing(df.columns, TARGET_MONTH_CANDIDATES)
    original_col = _first_existing(df.columns, TARGET_ORIGINAL_CANDIDATES)
    revised_col = _first_existing(df.columns, TARGET_REVISED_CANDIDATES)
    annual_col = _first_existing(df.columns, TARGET_ANNUAL_CANDIDATES)
    notes_col = _first_existing(df.columns, TARGET_NOTES_CANDIDATES)
    month_cols = {month: col for col in df.columns if (month := _month_column_number(col))}

    if year_col and month_col and original_col:
        score = 80
        score += 10 if revised_col else 0
        score += 5 if annual_col else 0
        return TargetSheetCandidate(sheet_name, header_row, "long", score, year_col, month_col, original_col, revised_col, annual_col, notes_col, {})

    if year_col and len(month_cols) >= 3:
        score = 70 + min(len(month_cols), 12)
        score += 5 if annual_col else 0
        return TargetSheetCandidate(sheet_name, header_row, "wide", score, year_col, None, None, revised_col, annual_col, notes_col, dict(sorted(month_cols.items())))

    return None


def analyze_target_workbook(excel_file) -> TargetWorkbookAnalysis:
    workbook = pd.ExcelFile(excel_file)
    candidates: list[TargetSheetCandidate] = []
    for sheet_name in workbook.sheet_names:
        for header_row in range(0, 10):
            df = _read_sheet(excel_file, sheet_name, header_row)
            if df is None:
                continue
            candidate = _score_candidate(df, sheet_name, header_row)
            if candidate:
                candidates.append(candidate)
    candidates = sorted(candidates, key=lambda item: item.score, reverse=True)
    return TargetWorkbookAnalysis(candidates=candidates, sheet_names=workbook.sheet_names)


def workbook_looks_like_sales_data(excel_file) -> bool:
    try:
        workbook = pd.ExcelFile(excel_file)
    except Exception:
        return False
    required = set(REQUIRED_COLUMNS)
    for sheet_name in workbook.sheet_names:
        try:
            preview = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, nrows=12)
        except Exception:
            continue
        for row_index in range(len(preview)):
            values = {str(value).strip() for value in preview.iloc[row_index].dropna().tolist()}
            if required.issubset(values):
                return True
    return False


def _clean_year(value: object) -> int | None:
    if pd.isna(value):
        return None
    try:
        year = int(float(str(value).strip()))
    except ValueError:
        return None
    return year if 1900 <= year <= 2200 else None


def normalize_targets(excel_file, candidate: TargetSheetCandidate) -> tuple[pd.DataFrame, dict[int, float]]:
    df = _read_sheet(excel_file, candidate.sheet_name, candidate.header_row)
    if df is None:
        raise ValueError("目标工作表为空或无法读取。")

    annual_targets: dict[int, float] = {}
    records: list[dict[str, object]] = []
    if candidate.layout == "long":
        for _, row in df.iterrows():
            year = _clean_year(row.get(candidate.year_column))
            month = parse_month(row.get(candidate.month_column))
            if year is None or month is None:
                continue
            original_target = pd.to_numeric(str(row.get(candidate.original_target_column, "")).replace(",", "").replace("£", ""), errors="coerce")
            if pd.isna(original_target):
                continue
            revised_target = pd.NA
            if candidate.revised_target_column:
                revised_target = pd.to_numeric(str(row.get(candidate.revised_target_column, "")).replace(",", "").replace("£", ""), errors="coerce")
            notes = row.get(candidate.notes_column) if candidate.notes_column else ""
            records.append(
                {
                    "Year": year,
                    "Month": month,
                    "Month Label": MONTH_LABELS[month],
                    "Original Target": float(original_target),
                    "Revised Target": float(revised_target) if pd.notna(revised_target) else pd.NA,
                    "Notes": "" if pd.isna(notes) else str(notes),
                }
            )
            if candidate.annual_target_column:
                annual_value = pd.to_numeric(str(row.get(candidate.annual_target_column, "")).replace(",", "").replace("£", ""), errors="coerce")
                if pd.notna(annual_value):
                    annual_targets[year] = float(annual_value)
    else:
        for _, row in df.iterrows():
            year = _clean_year(row.get(candidate.year_column))
            if year is None:
                continue
            if candidate.annual_target_column:
                annual_value = pd.to_numeric(str(row.get(candidate.annual_target_column, "")).replace(",", "").replace("£", ""), errors="coerce")
                if pd.notna(annual_value):
                    annual_targets[year] = float(annual_value)
            notes = row.get(candidate.notes_column) if candidate.notes_column else ""
            for month, col in candidate.month_columns.items():
                original_target = pd.to_numeric(str(row.get(col, "")).replace(",", "").replace("£", ""), errors="coerce")
                if pd.isna(original_target):
                    continue
                records.append(
                    {
                        "Year": year,
                        "Month": month,
                        "Month Label": MONTH_LABELS[month],
                        "Original Target": float(original_target),
                        "Revised Target": pd.NA,
                        "Notes": "" if pd.isna(notes) else str(notes),
                    }
                )

    target_df = pd.DataFrame.from_records(records)
    if target_df.empty:
        raise ValueError("未能从目标工作表中提取有效月份目标。")
    target_df = (
        target_df.sort_values(["Year", "Month"])
        .drop_duplicates(subset=["Year", "Month"], keep="last")
        .reset_index(drop=True)
    )
    target_df["Revised Target"] = target_df["Revised Target"].fillna(target_df["Original Target"])
    return target_df, annual_targets


def manual_candidate(
    sheet_name: str,
    header_row: int,
    layout: str,
    year_column: str,
    month_column: str | None,
    original_target_column: str | None,
    annual_target_column: str | None = None,
) -> TargetSheetCandidate:
    return TargetSheetCandidate(
        sheet_name=sheet_name,
        header_row=header_row,
        layout=layout,
        score=0,
        year_column=year_column,
        month_column=month_column,
        original_target_column=original_target_column,
        revised_target_column=None,
        annual_target_column=annual_target_column,
        notes_column=None,
        month_columns={},
    )


def read_sheet_columns(excel_file, sheet_name: str, header_row: int) -> list[str]:
    df = _read_sheet(excel_file, sheet_name, header_row)
    return [] if df is None else list(df.columns)
