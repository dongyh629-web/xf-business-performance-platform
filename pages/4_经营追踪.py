from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from app.auth import require_login
from app.config import DATE_BASIS_LABELS
from app.google_drive import MANUAL_SOURCE_LABEL, ensure_drive_data_loaded, render_data_source_sidebar
from app.target_metrics import (
    MONTH_LABELS,
    parse_xf_target_workbook,
    workbook_looks_like_sales_data,
)
from app.tracking_metrics import (
    analysis_year,
    average_allocation_table,
    build_monthly_tracking_table,
    build_product_group_amount_tracking,
    build_product_group_case_tracking,
)
from app.ui import inject_global_styles, money, percent, section_header, show_code_warning, show_context_summary, show_filters


def _format_percent(value: float | None) -> str:
    return "无基准" if value is None or pd.isna(value) else percent(float(value))


def _format_money_or_blank(value: float | None) -> str:
    if value is None or pd.isna(value):
        return ""
    return money(float(value))


def _safe_float(value: object, default: float = 0.0) -> float:
    number = pd.to_numeric(value, errors="coerce")
    return default if pd.isna(number) else float(number)


def _complete_target_editor_rows(targets: pd.DataFrame, target_year: int) -> pd.DataFrame:
    year_targets = targets[targets["Year"].astype("Int64").eq(target_year)].copy()
    records = []
    for month in range(1, 13):
        rows = year_targets[year_targets["Month"].astype(int).eq(month)]
        if rows.empty:
            original = 0.0
            revised = 0.0
            notes = ""
        else:
            row = rows.iloc[-1]
            original = _safe_float(row.get("Original Target", 0.0))
            revised = _safe_float(row.get("Revised Target", original), original)
            notes = "" if pd.isna(row.get("Notes", "")) else str(row.get("Notes", ""))
        records.append(
            {
                "Year": target_year,
                "Month": month,
                "Month Label": MONTH_LABELS[month],
                "Original Target": original,
                "Revised Target": revised,
                "Notes": notes,
            }
        )
    return pd.DataFrame.from_records(records)


def _target_years(targets: pd.DataFrame | None, fallback_year: int) -> list[int]:
    if targets is None or targets.empty or "Year" not in targets.columns:
        return [fallback_year]
    years = sorted(targets["Year"].dropna().astype(int).unique().tolist())
    return years or [fallback_year]


def _apply_revised_targets(targets: pd.DataFrame, target_year: int, edited: pd.DataFrame) -> pd.DataFrame:
    existing = targets[~targets["Year"].astype("Int64").eq(target_year)].copy() if not targets.empty else pd.DataFrame()
    revised_rows = []
    for _, row in edited.iterrows():
        month = int(row["Month"])
        original = _safe_float(row["Original Target"])
        revised = _safe_float(row["Revised Target"], original)
        revised_rows.append(
            {
                "Year": target_year,
                "Month": month,
                "Month Label": MONTH_LABELS[month],
                "Original Target": original,
                "Revised Target": revised,
                "Notes": "" if pd.isna(row.get("Notes", "")) else str(row.get("Notes", "")),
            }
        )
    return pd.concat([existing, pd.DataFrame.from_records(revised_rows)], ignore_index=True).sort_values(["Year", "Month"])


def _validate_target_rows(rows: pd.DataFrame) -> list[str]:
    errors = []
    months = pd.to_numeric(rows["Month"], errors="coerce").dropna().astype(int).tolist()
    if len(months) != len(set(months)):
        errors.append("月份不能重复。")
    missing = sorted(set(range(1, 13)) - set(months))
    if missing:
        errors.append(f"当前目标缺少月份：{', '.join(str(month) for month in missing)}。")
    for column in ["Original Target", "Revised Target"]:
        values = pd.to_numeric(rows[column], errors="coerce")
        if values.isna().any():
            errors.append(f"{column} 必须是数字。")
        if values.lt(0).any():
            errors.append(f"{column} 不允许为负数。")
    return errors


def _store_targets(
    target_df: pd.DataFrame,
    annual_targets: dict[int, float] | None = None,
    structure_label: str = "网页手动建立目标",
    amount_df: pd.DataFrame | None = None,
    case_df: pd.DataFrame | None = None,
) -> None:
    target_df = target_df.copy()
    target_df["Revised Target"] = pd.to_numeric(target_df["Revised Target"], errors="coerce").fillna(
        pd.to_numeric(target_df["Original Target"], errors="coerce")
    )
    st.session_state["target_data"] = target_df
    st.session_state["target_annual_targets"] = annual_targets or {
        int(year): float(group["Revised Target"].sum()) for year, group in target_df.groupby("Year")
    }
    st.session_state["target_structure_label"] = structure_label
    if amount_df is not None:
        st.session_state["target_amount_data"] = amount_df
    if case_df is not None:
        st.session_state["target_case_data"] = case_df


def _sync_home_targets(target_year: int, target_df: pd.DataFrame, anchor_month: int | None = None) -> None:
    year_targets = target_df[target_df["Year"].astype("Int64").eq(target_year)].copy()
    if year_targets.empty:
        return
    st.session_state["home_annual_target"] = float(pd.to_numeric(year_targets["Revised Target"], errors="coerce").fillna(0).sum())
    if anchor_month:
        month_rows = year_targets[year_targets["Month"].astype(int).eq(anchor_month)]
        if not month_rows.empty:
            st.session_state["home_monthly_target"] = _safe_float(month_rows.iloc[-1]["Revised Target"])


def _sales_cutoff_text(sales_df: pd.DataFrame | None) -> str:
    if sales_df is None or "Performance Date" not in sales_df.columns:
        return "无"
    dates = pd.to_datetime(sales_df["Performance Date"], errors="coerce").dropna()
    return "无" if dates.empty else str(dates.max().date())


def _target_years_text(targets: pd.DataFrame | None) -> str:
    if targets is None or targets.empty or "Year" not in targets.columns:
        return "无"
    years = sorted(targets["Year"].dropna().astype(int).unique().tolist())
    return ", ".join(str(year) for year in years) if years else "无"


def _show_data_status(sales_df: pd.DataFrame | None) -> None:
    target_df = st.session_state.get("target_data")
    status_cols = st.columns(2)
    with status_cols[0]:
        st.markdown("**销售数据**")
        if sales_df is None:
            st.caption("状态：未加载")
            st.caption("当前文件名：无")
            st.caption("数据截止日期：无")
        else:
            st.caption("状态：已加载")
            st.caption(f"来源：{st.session_state.get('data_source', '未知')}")
            st.caption(f"当前文件名：{st.session_state.get('source_file_name') or st.session_state.get('current_file_name') or '无'}")
            st.caption(f"Drive 最后修改时间：{st.session_state.get('sales_drive_modified_time', '无')}")
            st.caption(f"数据截止日期：{_sales_cutoff_text(sales_df)}")
    with status_cols[1]:
        st.markdown("**目标数据**")
        if target_df is None:
            st.caption("状态：未加载")
            st.caption("当前文件名：无")
            st.caption("已识别年份：无")
            st.caption("已识别工作表：无")
        else:
            st.caption("状态：已加载")
            st.caption(f"来源：{st.session_state.get('target_source', '当前会话')}")
            st.caption(f"当前文件名：{st.session_state.get('target_excel_name', '网页手动目标')}")
            st.caption(f"Drive 最后修改时间：{st.session_state.get('target_drive_modified_time', '无')}")
            st.caption(f"已识别年份：{_target_years_text(target_df)}")
            st.caption(f"已识别工作表：{st.session_state.get('target_structure_label', '网页手动建立目标')}")


def _display_money_table(df: pd.DataFrame, money_columns: list[str], percent_columns: list[str]) -> pd.DataFrame:
    display = df.copy()
    for column in money_columns:
        if column in display.columns:
            display[column] = display[column].map(_format_money_or_blank)
    for column in percent_columns:
        if column in display.columns:
            display[column] = display[column].map(_format_percent)
    return display


def _clean_excel_value(value: object) -> object:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and (pd.isna(value) or value in (float("inf"), float("-inf"))):
        return ""
    return value


def _append_dataframe_sheet(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title)
    clean_df = df.replace([float("inf"), float("-inf")], pd.NA).where(pd.notna(df), "")
    for row in dataframe_to_rows(clean_df, index=False, header=True):
        ws.append([_clean_excel_value(value) for value in row])
    ws.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for col_idx, column in enumerate(clean_df.columns, 1):
        letter = get_column_letter(col_idx)
        max_len = max([len(str(column))] + [len(str(value)) for value in clean_df[column].head(200).tolist()])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 34)
        lower = str(column).lower()
        for cell in ws[letter][1:]:
            if isinstance(cell.value, (int, float)):
                if any(token in lower for token in ["rate", "yoy", "完成率", "同比"]):
                    cell.number_format = "0.0%"
                elif any(token in lower for token in ["target", "sales", "gap", "目标", "销售", "缺口", "金额", "实际"]):
                    cell.number_format = '£#,##0'


def _report_notes(filtered: pd.DataFrame, target_name: str, target_year: int, summary) -> pd.DataFrame:
    basis = filtered.attrs.get("date_basis", st.session_state.get("date_basis", "Completed Date"))
    basis_label = DATE_BASIS_LABELS.get(basis, basis)
    date_range = filtered.attrs.get("date_range") or ("全部", "全部")
    customer_types = filtered.attrs.get("customer_types", [])
    product_groups = filtered.attrs.get("product_groups", [])
    return pd.DataFrame(
        [
            ("报告生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("销售数据来源文件", st.session_state.get("source_file_name") or st.session_state.get("current_file_name") or "无"),
            ("目标数据来源文件", target_name),
            ("数据截止日期", str(summary.analysis_date.date())),
            ("分析年度", target_year),
            ("Date Basis", basis_label),
            ("日期范围", f"{date_range[0]} 至 {date_range[1]}" if isinstance(date_range, tuple) else str(date_range)),
            ("客户类型筛选", "全部" if not customer_types else ", ".join(map(str, customer_types))),
            ("产品组筛选", "全部" if not product_groups else ", ".join(map(str, product_groups))),
            ("年度口径", "自然年：1月1日至12月31日"),
            ("当前月份说明", "当前进行中月份按分析截止日期统计，同比按同月同日进度比较。"),
            ("Gap公式", "Actual Sales - Revised Target；目标缺口=max(Revised Target - Actual Sales, 0)。"),
            ("数量口径说明", "实际箱数需要结合 SKU 箱规进一步计算，当前报告不输出虚假实际箱数。"),
        ],
        columns=["项目", "说明"],
    )


def _build_tracking_report(
    filtered: pd.DataFrame,
    target_year: int,
    target_df: pd.DataFrame,
    monthly_table: pd.DataFrame,
    summary,
    amount_table: pd.DataFrame,
    case_table: pd.DataFrame,
    allocation: pd.DataFrame,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    basis = filtered.attrs.get("date_basis", st.session_state.get("date_basis", "Completed Date"))
    summary_df = pd.DataFrame(
        [
            ("分析年度", target_year),
            ("数据截止日期", str(summary.analysis_date.date())),
            ("Date Basis", DATE_BASIS_LABELS.get(basis, basis)),
            ("年度原始目标", float(pd.to_numeric(target_df[target_df["Year"].astype("Int64").eq(target_year)]["Original Target"], errors="coerce").fillna(0).sum())),
            ("年度调整后目标", summary.annual_target),
            ("年度累计实际", summary.annual_actual),
            ("年度完成率", summary.annual_completion),
            ("年度累计同比", summary.annual_yoy),
            ("年度 Gap", summary.annual_actual - summary.annual_target),
            ("已结束月份正式缺口", summary.formal_shortfall),
            ("剩余月份数", summary.future_month_count),
        ],
        columns=["指标", "值"],
    )
    _append_dataframe_sheet(wb, "经营摘要", summary_df)
    _append_dataframe_sheet(wb, "月度追踪", monthly_table)
    _append_dataframe_sheet(wb, "系列金额追踪", amount_table)
    if case_table.empty:
        case_report = pd.DataFrame([{"说明": "实际箱数需要结合 SKU 箱规进一步计算，当前不输出虚假实际箱数。"}])
    else:
        case_report = case_table
    _append_dataframe_sheet(wb, "系列数量追踪", case_report)
    _append_dataframe_sheet(wb, "目标明细", target_df)
    allocation_report = allocation.copy()
    if allocation_report.empty:
        allocation_report = pd.DataFrame([{"说明": "当前没有未来月份可承接缺口。该数据为建议，不会自动覆盖 Revised Target。"}])
    else:
        allocation_report["说明"] = "该数据为建议，不会自动覆盖 Revised Target。"
    _append_dataframe_sheet(wb, "调整建议", allocation_report)
    _append_dataframe_sheet(wb, "报告说明", _report_notes(filtered, st.session_state.get("target_excel_name", "网页手动目标"), target_year, summary))

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


st.set_page_config(page_title="经营追踪", layout="wide")
inject_global_styles()
require_login("business_tracking")
st.title("经营追踪")
st.caption("Business Tracking")

ensure_drive_data_loaded()
render_data_source_sidebar(show_uploaders=False)

df = st.session_state.get("clean_data")

section_header("数据状态")
_show_data_status(df)

section_header("上传目标表 / Upload Targets Excel")
uploaded_target = st.file_uploader("上传目标表 / Upload Targets Excel", type=["xlsx"], key="target_excel_upload")
if uploaded_target is not None:
    target_bytes = uploaded_target.getvalue()
    try:
        parsed = parse_xf_target_workbook(BytesIO(target_bytes))
    except ValueError as exc:
        if workbook_looks_like_sales_data(BytesIO(target_bytes)):
            st.error("该文件看起来像销售明细，不像目标表。请使用左侧‘上传销售明细’入口。")
        else:
            st.error(str(exc))
    else:
        st.session_state["pending_target_excel_name"] = uploaded_target.name
        st.session_state["pending_target_parsed"] = parsed
        st.success("已识别目标文件，请确认后导入当前会话。")

pending = st.session_state.get("pending_target_parsed")
if pending is not None:
    cols = st.columns(4)
    cols[0].metric("识别年度", pending.target_year or "无")
    cols[1].metric("产品系列数量", pending.product_group_count)
    cols[2].metric("公司年度金额目标", _format_money_or_blank(pending.company_annual_amount_target))
    cols[3].metric("公司年度箱数目标", "" if pending.company_annual_case_target is None else f"{pending.company_annual_case_target:,.0f}")
    st.caption(f"文件名：{st.session_state.get('pending_target_excel_name')}")
    st.caption(f"金额目标工作表：{pending.amount_sheet} | 数量目标工作表：{pending.case_sheet}")
    st.dataframe(pending.amount_data.head(5), width="stretch", hide_index=True)
    if st.button("确认导入目标到当前会话", type="primary"):
        _store_targets(
            pending.company_targets,
            pending.annual_targets,
            pending.structure_label,
            pending.amount_data,
            pending.case_data,
        )
        st.session_state["target_excel_name"] = st.session_state.get("pending_target_excel_name", "目标 Excel")
        st.session_state["target_source"] = MANUAL_SOURCE_LABEL
        st.session_state["target_source_type"] = "manual"
        st.session_state.pop("target_drive_modified_time", None)
        if pending.target_year:
            _sync_home_targets(pending.target_year, pending.company_targets)
        st.success("目标数据已导入当前会话。")
        st.rerun()

target_df = st.session_state.get("target_data")
current_year = analysis_year(df) if df is not None else None
fallback_year = current_year or datetime.now().year

section_header("目标管理 / Target Management")
st.caption("Original Target 可由 Google Drive 或手动上传导入；Revised Target 仅在当前会话中调整，不会写回 Google Drive。")
available_years = _target_years(target_df, fallback_year)
default_year_index = available_years.index(current_year) if current_year in available_years else 0
target_year = st.selectbox("目标年度", available_years, index=default_year_index)

base_targets = target_df if target_df is not None else pd.DataFrame(columns=["Year", "Month", "Month Label", "Original Target", "Revised Target", "Notes"])
annual_input = st.number_input(
    "年度销售目标",
    min_value=0.0,
    value=float(pd.to_numeric(_complete_target_editor_rows(base_targets, target_year)["Revised Target"], errors="coerce").fillna(0).sum()),
    step=50000.0,
    format="%.0f",
)
editor_rows = _complete_target_editor_rows(base_targets, target_year)
if st.button("平均分配年度目标"):
    average_target = annual_input / 12 if annual_input else 0.0
    editor_rows["Original Target"] = average_target
    editor_rows["Revised Target"] = average_target

edited_targets = st.data_editor(
    editor_rows,
    width="stretch",
    hide_index=True,
    disabled=["Year", "Month", "Month Label", "Original Target"],
    column_config={
        "Year": st.column_config.NumberColumn("Year", format="%d"),
        "Month": st.column_config.NumberColumn("Month", format="%d"),
        "Month Label": st.column_config.TextColumn("Month"),
        "Original Target": st.column_config.NumberColumn("Original Target", format="£%.0f", min_value=0),
        "Revised Target": st.column_config.NumberColumn("Revised Target", format="£%.0f", min_value=0),
        "Notes": st.column_config.TextColumn("Notes"),
    },
    key=f"target_editor_{target_year}",
)

button_cols = st.columns(2)
with button_cols[0]:
    if st.button("保存目标到当前会话", type="primary"):
        errors = _validate_target_rows(edited_targets)
        if errors:
            for error in errors:
                st.error(error)
        else:
            next_targets = _apply_revised_targets(base_targets, target_year, edited_targets)
            _store_targets(
                next_targets,
                structure_label=st.session_state.get("target_structure_label", "网页手动建立目标"),
                amount_df=st.session_state.get("target_amount_data"),
                case_df=st.session_state.get("target_case_data"),
            )
            st.session_state["target_source"] = "当前会话手动调整"
            st.session_state["target_source_type"] = "manual"
            _sync_home_targets(target_year, next_targets)
            diff = float(pd.to_numeric(edited_targets["Revised Target"], errors="coerce").fillna(0).sum()) - annual_input
            if abs(diff) > 0.01:
                st.warning(f"12个月 Revised Target 合计与输入年度目标差额为 {money(diff)}。")
            st.success("目标已保存到当前会话。")
            st.rerun()
with button_cols[1]:
    if st.button("恢复为原始目标"):
        restored = edited_targets.copy()
        restored["Revised Target"] = restored["Original Target"]
        next_targets = _apply_revised_targets(base_targets, target_year, restored)
        _store_targets(
            next_targets,
            structure_label=st.session_state.get("target_structure_label", "网页手动建立目标"),
            amount_df=st.session_state.get("target_amount_data"),
            case_df=st.session_state.get("target_case_data"),
        )
        st.session_state["target_source"] = "当前会话手动调整"
        st.session_state["target_source_type"] = "manual"
        _sync_home_targets(target_year, next_targets)
        st.success("已恢复为原始目标。")
        st.rerun()

target_df = st.session_state.get("target_data")
if target_df is None:
    st.info("请上传目标 Excel，或在上方手动建立目标，以查看月度目标、完成率、Gap 和年度追踪。")

if df is None:
    st.info("请使用左侧‘上传销售明细 / Upload Unleashed Sales Data’入口上传销售明细，以生成实际销售、完成率、Gap 和年度追踪。")
    st.stop()

filtered = show_filters(df, "tracking")
show_code_warning(filtered)
show_context_summary(filtered)

current_year = analysis_year(filtered)
if current_year is None:
    st.info("当前筛选结果没有有效 Performance Date，无法生成经营追踪。")
    st.stop()
st.caption(f"当前分析年度来自所选 Date Basis 下最大 Performance Date：{current_year}。")

target_df = st.session_state.get("target_data")
if target_df is None:
    st.stop()

monthly_table, summary = build_monthly_tracking_table(filtered, target_df, target_year)
_sync_home_targets(target_year, target_df, summary.analysis_date.month)

annual_target_from_file = st.session_state.get("target_annual_targets", {}).get(target_year)
if annual_target_from_file is not None:
    diff = summary.annual_target - annual_target_from_file
    if abs(diff) > 0.01:
        st.warning(f"目标表年度总目标为 {money(annual_target_from_file)}，12个月调整后目标合计为 {money(summary.annual_target)}，差额 {money(diff)}。")

section_header("年度追踪")
kpi_cols = st.columns(5)
kpi_cols[0].metric("年度目标", money(summary.annual_target))
kpi_cols[1].metric("年度累计实际", money(summary.annual_actual))
kpi_cols[2].metric("年度累计完成率", _format_percent(summary.annual_completion))
kpi_cols[3].metric("年度累计同比", _format_percent(summary.annual_yoy))
kpi_cols[4].metric("年度目标缺口", money(summary.annual_target_shortfall))

section_header("月度目标追踪")
display_table = _display_money_table(
    monthly_table,
    ["原始目标", "调整后目标", "实际销售", "去年同期", "Gap", "目标缺口", "累计实际", "累计目标"],
    ["完成率", "同比", "累计完成率"],
)
st.dataframe(
    display_table[
        [
            "月份",
            "原始目标",
            "调整后目标",
            "实际销售",
            "完成率",
            "去年同期",
            "同比",
            "Gap",
            "累计实际",
            "累计目标",
            "累计完成率",
            "状态",
        ]
    ],
    width="stretch",
    hide_index=True,
)

amount_targets = st.session_state.get("target_amount_data")
amount_table = build_product_group_amount_tracking(filtered, amount_targets, target_year) if amount_targets is not None else pd.DataFrame()
if not amount_table.empty:
    section_header("系列金额追踪")
    amount_display = _display_money_table(
        amount_table,
        ["Original Amount Target", "Revised Amount Target", "Actual Sales Amount", "Amount Gap", "Previous Year Actual"],
        ["Amount Completion Rate", "YoY"],
    )
    st.dataframe(amount_display, width="stretch", hide_index=True)

case_targets = st.session_state.get("target_case_data")
case_table = build_product_group_case_tracking(case_targets, target_year) if case_targets is not None else pd.DataFrame()
if not case_table.empty:
    section_header("系列数量目标")
    st.caption("实际箱数需要结合 SKU 箱规进一步计算，当前不使用订单数或销售行数代替。")
    st.dataframe(case_table, width="stretch", hide_index=True)

section_header("后续月份平均分摊建议")
st.caption("正式缺口只来自已结束月份的未完成部分；当前进行中月份暂不计入再分配，未来月份作为缺口承接月份。")
st.metric("待补正式缺口", money(summary.formal_shortfall))
allocation = average_allocation_table(monthly_table, summary.average_shortfall_allocation)
if allocation.empty:
    st.info("当前没有未来月份可承接缺口。")
else:
    allocation_display = _display_money_table(allocation, ["当前调整后目标", "平均追加目标", "建议调整后目标"], [])
    st.dataframe(allocation_display, width="stretch", hide_index=True)

section_header("下载经营追踪报告")
try:
    report_bytes = _build_tracking_report(filtered, target_year, target_df, monthly_table, summary, amount_table, case_table, allocation)
    load_workbook(BytesIO(report_bytes), read_only=True)
except Exception as exc:
    st.error(f"经营追踪报告生成失败：{exc}")
else:
    st.download_button(
        "下载经营追踪报告 / Download Tracking Report",
        report_bytes,
        file_name=f"XF_Business_Tracking_{summary.analysis_date.date()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
